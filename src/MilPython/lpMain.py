from scipy.sparse import vstack
import gurobipy as gp
import numpy as np
from .lpObject import LPObject
from .lpStateVar import LPStateVar,LPStateVar_timedep,LPStateVar_add
from .lpInputdata import LPInputdata
from scipy.sparse import coo_matrix, csc_matrix, vstack
from .tools import Solver,Obj

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sympy as sp
import tkinter as tk
from tkinter import font as tkfont
from tkinter import filedialog
from tkinter import ttk
import tempfile

class LPMain:
    '''
    (Abstract) main class for running the linear optimization
    Here the equation systems are prepared for optimization and the optimization is executed
    '''
    def __init__(self,inputdata:LPInputdata):
        """This is the init-fun of the abstract class LPMain. This code has to be run by inheriting class by >LPMain().__init__(self,inputdata)"""        
        if self.__class__.__name__ == 'LPMain':
            raise Exception('This is an abstract class. Please only instantiate objects of the inheriting class.')
        self.Aeq = None
        self.beq = []
        self.senses = []
        self.inputdata = inputdata
        self.make_stateVarLst()
        self.def_pos()
        self.def_bounds()
        self.def_vtypes()
        self.def_eqs()
        self.init_targetfun()
        self.def_targetfun()
    
    def def_eqs(self):
        """Defines the equation system. Calls the def_equation function for all LPObjects in self.obj_lst and extends Aeq,beq and senses by the equations of the objects"""        
        self.obj_lst[0].def_equations()
        self.Aeq,self.beq,self.senses = self.obj_lst[0].return_eqs()
        for obj in self.obj_lst[1:]:
            obj.def_equations()
            self.extend_matrices(obj.return_eqs())
    
    def extend_matrices(self,eq_lst):
        '''Appends equations from other classes to the equation system of the LPMain object'''
        self.Aeq = vstack([self.Aeq,eq_lst[0]])
        self.beq.extend(eq_lst[1])
        self.senses.extend(eq_lst[2])
    
    def make_stateVarLst(self):
        '''
        Creates lists containing all status variables of all objects belonging to the system.
        Divided into time-dependent variables and additional variables
        '''
        self.stateVars:list[LPStateVar]=[]
        for obj in self.obj_lst:
            self.stateVars.extend(obj.getStateVars())
        self.stateVars_timedep = [var for var in self.stateVars if isinstance(var,LPStateVar_timedep)]
        self.stateVars_add = [var for var in self.stateVars if isinstance(var,LPStateVar_add)]
           
    def def_pos(self):
        '''
        Defines the positions of all state variables within the Aeq matrix.
        - For time-dependent variables, the position of the variable is saved for the first time step
        - Additional variables are at the end of the list
        '''
        idx_pos=0
        for var in self.stateVars_timedep:
            var.pos = idx_pos
            idx_pos += 1
        idx_pos = len(self.stateVars_timedep)*self.inputdata.steps
        for var in self.stateVars_add:
            var.pos = idx_pos
            idx_pos += 1        
        self.inputdata.num_vars=idx_pos
        self.inputdata.num_vars_timedep=len(self.stateVars_timedep)

    def def_bounds(self):
        '''
        Creates lists containing the upper and lower limits of all state variables.
        The order corresponds to the positions assigned to the variables
        '''
        num_vars_timedep = len(self.stateVars_timedep)*self.inputdata.steps
        num_vars_add = len(self.stateVars_add)
        num_vars = num_vars_timedep + num_vars_add
        self.lb=np.zeros(num_vars)
        self.ub=np.zeros(num_vars)
        
        lb_timedep = []
        ub_timedep = []
        for var in self.stateVars_timedep:
            lb_timedep.append(var.lb)   
            ub_timedep.append(var.ub)   
        self.lb[0:num_vars_timedep] = lb_timedep*self.inputdata.steps
        self.ub[0:num_vars_timedep] = ub_timedep*self.inputdata.steps
                
        lb_add=[]
        ub_add=[]
        for var in self.stateVars_add:
            lb_add.append(var.lb)
            ub_add.append(var.ub)
        self.lb[num_vars_timedep:] = lb_add
        self.ub[num_vars_timedep:] = ub_add
    
    def def_vtypes(self):
        '''
        Creates lists containing the variable type of all state variables.
        The order corresponds to the positions assigned to the variables
        '''
        num_vars_timedep = len(self.stateVars_timedep)*self.inputdata.steps
        num_vars_add = len(self.stateVars_add)
        num_vars = num_vars_timedep + num_vars_add
        self.vtypes = []
        
        vtypes_timedep = []
        for var in self.stateVars_timedep:
            vtypes_timedep.append(var.vtype)   
        self.vtypes.extend(vtypes_timedep*self.inputdata.steps)                
        vtypes_add=[]
        for var in self.stateVars_add:
            vtypes_add.append(var.vtype)
        self.vtypes[num_vars_timedep:] = vtypes_add
        
    def init_targetfun(self):
        '''Inialization of the target functions with a zero vector'''
        self.f = np.zeros(self.inputdata.num_vars)
    
    def def_targetfun(self):
        pass
    
    def add_var_targetfun(self,var:LPStateVar,value,step=0):
        '''
        Adds a variable to the target function
        To do this, the StateVar, the desired time step and the weighting for the target function must be transferred
        For additional variables: don't add a variable for step
        '''
        self.f[var.pos+step*len(self.stateVars_timedep)]=value
        
        
    def optimize(self,mipGap=0.00,solver:Solver=Solver.GUROBI,objective:Obj=Obj.MINIMIZE):
        '''Performs the linear optimization of the system of equations set up'''
        if solver == Solver.GUROBI:
            x=self.solver_gurobi(mipGap,objective)
        elif solver == Solver.SCIPY:
            x=self.solver_scipy(mipGap,objective)
        elif solver == Solver.CPLEX:
            x=self.solver_cplex(mipGap,objective)
        else:
            raise Exception('This Solver is not implemented')
        self.assign_results(x)
    
    def assign_results(self,x):
        '''Assigns the results of the result vector x to the state variables'''
        self.x = x
        num_vars_timedep = len(self.stateVars_timedep)
        for var in self.stateVars_timedep:
            var.result = x[:num_vars_timedep*self.inputdata.steps][var.pos::num_vars_timedep]
        for var in self.stateVars_add:
            var.result = x[var.pos]
    
# %% Funktion Solver
    def solver_gurobi(self,mipGab,objective):
        '''
        The solver_gurobi function transfers the optimization model to the Gurobi solver, performs the optimization and returns the result.
        Aeq, beq, senses: Matrix or vector of equations with the comparison operator of each equation
        ctype, lb, ub: Type and upper and lower limits of the variables
        f: target function
        '''
        # Transfer the optimization problem to the Gurobi API.
        # Create an empty problem
        problem = gp.Model()
        if self.inputdata.verbose == False:
            problem.setParam('LogToConsole',0)
        # add variables
        x = problem.addMVar(shape=self.inputdata.num_vars,lb=self.lb,ub=self.ub,vtype=self.vtypes)
        # x = problem.addMVar(shape=self.inputdata.num_vars,lb=self.lb,ub=self.ub,vtype=['C' for _ in range(self.inputdata.num_vars)])
        # Pass target function
        
        if objective == Obj.MINIMIZE:
            problem.setObjective(self.f @ x, gp.GRB.MINIMIZE)    
        else:
            problem.setObjective(self.f @ x, gp.GRB.MAXIMIZE)    
        # pass equations
        problem.addMConstr(self.Aeq.tocsr(), x, self.senses, self.beq)
        # optimize problem
        problem.setParam('MIPGap', mipGab)  # Percentage distance to the optimum solution
        problem.optimize()
        x = x.X
        return x
    
    def solver_scipy(self,mipGap,objective):
        '''
        The solver_scipy function transfers the optimization model to the scipy milp solver, performs the optimization and returns the result.
        Note: This solver is free but very slow
        
        Aeq, beq, senses: Matrix or vector of equations with the comparison operator of each equation
        ctype, lb, ub: Type and upper and lower limits of the variables
        f: target function
        '''
        if objective == Obj.MINIMIZE:
            pass   
        else:
            raise Exception('The scipy-Solver only allows minimizing')  
        b_l=[]
        b_u=[]
        for idx,sense in enumerate(self.senses):
            if sense == 'E' or sense == '=':
                b_l.append(self.beq[idx])
                b_u.append(self.beq[idx])
            elif sense == '<':
                b_l.append(-np.inf)
                b_u.append(self.beq[idx])
            elif sense == '>':
                b_l.append(self.beq[idx])
                b_u.append(np.inf)
            else:
                raise Exception(f'Unknown Sense {sense}')
        # vtypes / integrality
        integrality=[]
        for idx,vtype in enumerate(self.vtypes):
            if vtype == 'C':
                integrality.append(0)
            elif vtype == 'I':
                integrality.append(1)
            elif vtype == 'S':
                integrality.append(2)
            elif vtype == 'N':
                integrality.append(3)#
            elif vtype == 'B':#
                integrality.append(1)
                self.lb[idx]=0
                self.ub[idx]=1
            else:
                print('unknown vtype')
        from scipy.optimize import LinearConstraint
        constraints = LinearConstraint(self.Aeq,b_l,b_u)
        # %%
        from scipy.optimize import milp
        res = milp(c=self.f,constraints=constraints,integrality=integrality,options={'mip_rel_gap':mipGap})
        return res.x

    def solver_cplex(self,mipgap,objective):
        '''
        The solver_cplex() function transfers the optimization model to the cplex solver, performs the optimization and returns the result.
        Aeq, beq, senses: Matrix or vector of equations with the comparison operator of each equation
        ctype, lb, ub: Type and upper and lower limits of the variables
        f: target function
        '''
        import cplex
        # nbew empty problem
        problem = cplex.Cplex()

        if self.inputdata.verbose == False:
            problem.set_log_stream(None)
            problem.set_results_stream(None)
        # Adding variables
        problem.variables.add(names=['x' + str(i) for i in range(self.inputdata.num_vars)])

        # formatting the variable-type-list to fit cplex
        types = [(i, problem.variables.type.continuous) if vtype_i == 'C' 
                    else (i, problem.variables.type.binary) if vtype_i == 'B'
                    else (i, problem.variables.type.semi_continuous) if vtype_i == 'S'
                    else (i, problem.variables.type.semi_integer) if vtype_i == 'N'
                    else (i, problem.variables.type.integer) for i, vtype_i in enumerate(self.vtypes)]
        problem.variables.set_types(types)

        # Setting lower and upper bounds
        problem.variables.set_lower_bounds(list(enumerate(self.lb)))
        problem.variables.set_upper_bounds(list(enumerate(self.ub)))

        # set targetfunction
        problem.objective.set_linear(list(enumerate(self.f)))
        # set objective
        if objective == Obj.MINIMIZE:
            problem.objective.set_sense(problem.objective.sense.minimize)    
        else:
            problem.objective.set_sense(problem.objective.sense.maximize)    

        # prepare equations
        Aeq_rows = self.Aeq.row.tolist()
        Aeq_cols = self.Aeq.col.tolist()
        Aeq_vals = self.Aeq.data.tolist()
        beq_rows = [i for i in range(len(self.beq))]
        beq_vals = np.array(self.beq,dtype=float)

        # cplex need var-names
        problem.linear_constraints.add(names=['c' + str(i) for i in range(np.shape(self.Aeq)[0])])
        # setting Aeq
        problem.linear_constraints.set_coefficients(zip(Aeq_rows, Aeq_cols, Aeq_vals))
        # setting beq
        problem.linear_constraints.set_rhs(zip(beq_rows, beq_vals))

        # preparing and setting senses
        senses =[]
        for sense in self.senses:
            if sense == '<':
                senses.append('L')
            elif sense == '=':
                senses.append('E')
            elif sense == '>':
                senses.append('G')
            else:
                senses.append(sense)
        problem.linear_constraints.set_senses(list(enumerate(senses)))
            
        # setting mipgap
        problem.parameters.mip.tolerances.mipgap.set(float(mipgap))

        del Aeq_rows, Aeq_cols, Aeq_vals, beq_rows, beq_vals

        # Solver
        problem.solve()

        # Returning result vector
        x = np.array(problem.solution.get_values())
        return x
            

    def results_to_excel(self,path):
        '''
        Exports the results of the LP model to an Excel file. 
        Iterates through each object in 'obj_lst', extracts state variables and their results, 
        and writes them to separate sheets in an Excel file.

        Parameters:
        path (str): The file path where the Excel file should be saved.
        '''
        if not path.endswith('.xlsx'):
            path+='.xlsx'
        import pandas as pd
        with pd.ExcelWriter(path,engine='openpyxl') as writer:
            for obj in self.obj_lst:
                if obj.name=='':
                    name=type(obj).__name__
                else:
                    name = obj.name
                df = pd.DataFrame(columns=['Name','Unit','Type','Min','Max','Timedependet'])
                # df.columns = ['Name','Unit','Type','Min','Max','Timedependet']
                for var in obj.stateVar_lst:
                    td = type(var) == LPStateVar_timedep
                    df = pd.concat([df,pd.DataFrame.from_records([{'Name':var.name,'Unit':var.unit,'Type':var.vtype,'Min':var.lb,'Max':var.ub,'Timedependet':td}])])
                if len(df)>0:
                    df.to_excel(writer,sheet_name=f'{name} - Variables',index=False)
                # Result of timedependent variables
                df = pd.DataFrame()
                df['t']=range(self.inputdata.steps)
                for var in obj.stateVar_lst:
                    df[var.name]=var.result
                if df.shape[1]>1:
                    df.to_excel(writer,sheet_name=f'{name} - Results',index=False)
        self.__optimize_excel_col_width(path)
        
    def __optimize_excel_col_width(self,path):
        '''
        optimizes the width of the excel columns so the full title of each column can be read
        '''
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active

        # Funktion zum Berechnen der optimalen Breite für jede Spalte
        def get_optimal_width(column_values):
            max_length = 0
            for value in column_values:
                max_length = max(max_length, len(str(value)))
            return max_length + 2  # Zusätzlicher Platz für Puffer

        # Schleife durch jede Spalte im DataFrame
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Spalten als Liste holen
            columns = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            # Schleife durch jede Spalte im aktuellen Blatt
            for col_idx, column in enumerate(columns, start=1):  # Spaltenindex beginnt bei 1 in Excel
                col_letter = ws.cell(row=1, column=col_idx).column_letter  # Spaltenbuchstabe
                column_values = [cell.value for cell in ws[col_letter]]  # Alle Werte der Spalte
                optimal_width = get_optimal_width(column_values)
                ws.column_dimensions[col_letter].width = optimal_width

        # Speichern der angepassten Excel-Datei
        wb.save(path)

    def show_lp_system(self,window_size="1225x850",width=500):
        '''
        Displays a graphical user interface (GUI) for visualizing the Linear Programming (LP) system. 
        Sets up a Tkinter window with a notebook interface, creating tabs for each object in 'obj_lst'. 
        There is no auto formatting yet. This function is still in an early state.
        If the equation is to wide for hte window, increase the width
        '''
        root = tk.Tk()
        root.title("MILP System")
        root.geometry(window_size)  # Feste Fenstergröße einstellen

        # Notebook-Widget erstellen
        notebook = ttk.Notebook(root)
        notebook.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        for obj in self.obj_lst:
            tab = ScrollableFrame(notebook)
            notebook.add(tab, text=obj.name if obj.name else type(obj).__name__)
            frame = tab.scrollable_frame

            heading_font = tkfont.Font(size=16, weight="bold")

            if len(obj.stateVar_lst) > 0:
                # Display state variables in table
                label = ttk.Label(frame, text="State Variables", anchor="center", font=heading_font)
                label.pack(padx=width,pady=10, fill=tk.BOTH)
                
                tree = ttk.Treeview(frame, columns=('Name', 'Unit', 'Type', 'Min', 'Max', 'Timedep'), show='headings')
                tree.heading('Name', text='Name')
                tree.heading('Unit', text='Unit')
                tree.heading('Type', text='Type')
                tree.heading('Min', text='Min')
                tree.heading('Max', text='Max')
                tree.heading('Timedep', text='Timedependent')
                tree.pack(fill=tk.BOTH, expand=True)

                for var in obj.stateVar_lst:
                    var.timedep = isinstance(var, LPStateVar_timedep)

                sorted_state_vars = sorted(obj.stateVar_lst, key=lambda x: (x.timedep, x.name))
                for var in sorted_state_vars:
                    tree.insert("", "end", values=(var.name, var.unit, var.vtype, var.lb, var.ub, var.timedep))
            else:
                label = ttk.Label(frame, text="No State Variables", anchor="center", font=heading_font)
                label.pack(padx=width,pady=10, fill=tk.BOTH)

            label = ttk.Label(frame, text="Equations", anchor="center", font=heading_font)
            label.pack(pady=10, fill=tk.BOTH)

            eq_lst = obj.return_grouped_eqs()
            for eqs in eq_lst:
                label = ttk.Label(frame, text=eqs[0].description, anchor="center")
                label.pack(pady=10, fill=tk.BOTH)
                timesteps_lst = []
                for eq in eqs:
                    first_timeindex = None
                    for var_info in eq.var_lst:
                        if len(var_info) == 3:
                            _, _, t_step = var_info
                        if first_timeindex is None:
                            first_timeindex = t_step
                    timesteps_lst.append(first_timeindex)
                timesteps_str = f'für t = [{self.__summarize_intervals(timesteps_lst)}]'
                label = ttk.Label(frame, text=timesteps_str, anchor="center")
                label.pack(pady=10, fill=tk.BOTH)
                fig = self.__create_figure_for_display(eq)
                canvas = FigureCanvasTkAgg(fig, master=frame)
                canvas.draw()
                canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True) 

        root.update_idletasks()

        # Ensure window is on top
        root.attributes('-topmost', 1)
        root.attributes('-topmost', 0)  # Reset the topmost attribute
        root.mainloop()
        
    def __summarize_intervals(self,lst):
        output = []
        i = 0
        while i < len(lst):
            start = lst[i]
            while i+1 < len(lst) and lst[i+1] == lst[i]+1:
                i += 1
            end = lst[i]
            if start == end:
                output.append(str(start))
            else:
                output.append(f'{start}...{end}')
            i += 1
        return ', '.join(output)
    
    def __create_figure_for_display(self,eq):
        fig, ax = plt.subplots(figsize=(1.5,  0.5))  # Anpassen der Höhe an die Anzahl der Gleichungen #!anpassen
        ax.axis('off')

        y_position = 0.4
        fig.text(0.5, y_position, f"${sp.latex(self.__convert_to_sympy_equation(eq))}$", fontsize=12, ha='center', va='center')    

        return fig

    def __convert_to_sympy_equation(self,equation):
        # Initialize the left-hand side as 0
        lhs = 0
        first_timeindex=None
        for var_info in equation.var_lst:
            # Extract variable, factor, and time step
            if len(var_info) == 3:
                var, factor, t_step = var_info
                if first_timeindex==None:
                    first_timeindex=t_step
                if t_step==first_timeindex:
                    t_step='t'
                else:
                    t_dif=t_step-first_timeindex
                    if t_dif>0:
                        t_step=f't+{t_dif}'
                    else:
                        t_step=f't{t_dif}'
                # Create the variable with time index
                var_sympy = sp.Symbol(self.format_string(f"{var.name}_{{{t_step}}}"))
            else:
                var.name, factor = var_info
                # Create the variable without time index
                var_sympy = sp.Symbol(self.format_string(var.name))
            
            # Add term to the left-hand side
            if factor==1:
                lhs+=var_sympy
            elif factor ==-1:
                lhs+= -var_sympy
            else:
                lhs += self.round_scientific(factor) * var_sympy

        # Right-hand side
        rhs = equation.b

        # Create the sympy equation based on sense
        if equation.sense == 'E' or equation.sense=='e' or equation.sense=='=':
            return sp.Eq(lhs, rhs)
        elif equation.sense == '<':
            return sp.LessThan(lhs, rhs)
        elif equation.sense == '>':
            return sp.GreaterThan(lhs, rhs)
        else:
            raise ValueError("Unknown sense symbol. Use 'e' for equals, '<' for less than or equal, or '>' for greater than or equal.")
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self)
        self.scrollable_frame = ttk.Frame(self.canvas)

        # Vertikaler Scrollbar
        v_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        # Horizontaler Scrollbar
        h_scrollbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        self.bind("<Configure>", self.on_frame_configure)

        # Bind mouse wheel events for scrolling on the canvas only
        self.bind_mousewheel_events()

    def on_frame_configure(self, event):
        self.canvas.configure(width=event.width, height=event.height)
        self.canvas.configure(scrollregion=self.canvas.bbox('all'))

    def bind_mousewheel_events(self):
        self.canvas.bind("<Enter>", self._bind_to_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_shift_mousewheel(self, event):
        self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_to_mousewheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)

    def _unbind_mousewheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Shift-MouseWheel>")
    
    